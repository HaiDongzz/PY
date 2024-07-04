#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/3/8 15:39
# @Author  : HD
# @File    : Run_test.py.py
# @Description :

import openpyxl

from Api_request import sent_get_request, sent_post_request


# 读取测试用例
def read_test_cases(test_cases_path):
    wb = openpyxl.load_workbook(test_cases_path)
    if test_cases_path == None:
        pass
    try:
        wb = openpyxl.load_workbook(test_cases_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")

    sheet = wb.active
    test_cases = []
    for row in sheet.iter_rows(values_only=True):
        test_cases.append(row)
    # 跳过表头
    return test_cases[1:]


# 执行用例
def run_test(test_cases, request_func):
    for test_case in test_cases:
        url = test_case[0]
        expected_result = test_case[1]
        response = request_func(url)
        actual_result = response.get("result")
        assert actual_result == expected_result, f"TEST CASE FAILED FOR URL:{url}"


# 测试get请求
get_test_cases = read_test_cases("Test_case/test_case_get.xlsx")
run_test(get_test_cases, sent_get_request)

# 测试post请求
post_test_cases = read_test_cases("Test_case/test_case_post.xlsx")
run_test(post_test_cases, lambda url: sent_post_request(url, {"key": "value"}))
