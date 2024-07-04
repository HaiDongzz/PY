#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/11/20 14:30
# @Author  : HD
# @File    : Eam_Api_Test_Auto.py
# @Description :

import logging
import unittest

import openpyxl
import requests


class InterfaceTestCase(unittest.TestCase):
    def setUp(self):
        # 在测试之前的准备工作，例如初始化接口客户端
        self.base_url = "http://192.168.30.153:8002/eam"
        self.header = {
            'Authorization': 'eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJhZG1pbiIsInN1YiI6Iuezu-'
                             'e7n-euoeeQhuWRmCIsIkNVUlJFTlRfVVNFUl9JTkZPIjoie1wiY29tcGFueUF'
                             'yZWFJZFwiOlwiQVwiLFwiY29tcGFueUlkXCI6XCJBXCIsXCJjb21wYW55TmFtZV'
                             'wiOlwi6ZuG5Zui5pys57qnXCIsXCJlbWFpbFwiOlwiMDAwMDE3MzRAd3ptdHIuY29t'
                             'XCIsXCJtb2JpbGVcIjpcIjEzNzU4NDk2NTQ2XCIsXCJvZmZpY2VBcmVhSWRcIjpcIkEwM'
                             'lwiLFwib2ZmaWNlSWRcIjpcIkEwMlwiLFwib2ZmaWNlTmFtZVwiOlwi5Yqe5YWs5a6kXCIs'
                             'XCJwZXJzb25JZFwiOlwiYWRtaW5cIixcInBlcnNvbk5hbWVcIjpcIuezu-e7n-'
                             'euoeeQhuWRmFwiLFwicGVyc29uTm9cIjpcImFkbWluXCIsXCJwaG9uZVwiOlwiMDU3Ny04OT'
                             'cyNzAwM1wifSIsImlhdCI6MTcwMTM5MTU0MywiZXhwIjoxNzAxOTk'
                             '2MzQzfQ.nNj7LQ8m1oANG3jo8AiRy8CID-Xo-0Zc2JAppPiR_aY'
        }
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        handler = logging.StreamHandler()
        formatter = logging.Formatter('%(asctime)s' + '%(message)')
        handler.setFormatter(formatter)
        self.logger.addHandler(handler)

    def tearDown(self):
        # 在测试之后的清理工作，例如关闭接口客户端
        # 如果在setUp方法中初始化了接口客户端对象，需要在tearDown方法中将其关闭或销毁，释放资源
        self.close()  # 以client为例，具体方法根据实际情况调用
        pass

    '''        清理测试数据（示例：删除数据库中的测试数据）
        在测试过程中创建了测试数据，例如在数据库中插入了测试数据，可以在tearDown方法中删除这些测试数据
        cleanup_test_data()'''

    def test_interface(self):
        excel_file = 'D:\TEST-project\铁投内部系统项目\设备维修维护管理系统\测试文档\EAM接口测试用例.xlsx'  # Excel文件路径
        sheet_name = 'Sheet1'  # 表格名称

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):  # 从第2行开始，第1行为标题
            interface_name = sheet.cell(row=row, column=1).value  # 接口名称
            input_data = sheet.cell(row=row, column=4).value  # 入参
            input_url = sheet.cell(row=row, column=3).value  # 获取接口url
            expected_output = sheet.cell(row=row, column=5).value  # 期望输出

            # 转化预期结果为json
            if expected_output is None:
                expected_output == '{"code": 0, "data": "null", "message": "success", "referInfo": "null"}'

            # 调用接口进行测试，并获取实际输出
            actual_output = self.call_interface(interface_name, input_data, input_url)

            # 断言实际输出与期望输出是否一致
            try:
                self.assertEqual(actual_output, expected_output, f"接口 {interface_name} 测试失败")
                self.logger.info(f"{interface_name}测试通过；请求参数：{input_data};断言结果：{actual_output}")

            except AssertionError as e:
                self.logger.error(f"接口{interface_name}测试失败，请求参数:{input_data},错误信息：{e}")

                # assert actual_output == expected_output, "实际输出与期望输出不符"

    def call_interface(self, interface_name, input_data, input_url):
        # 调用接口进行测试，并返回实际输出
        # 根据实际情况，使用相应的库或工具调用接口
        if interface_name == '{interface_name}':
            url = f"{self.base_url}{input_url}"
            response = requests.post(url, headers=self.header, json=input_data)
            return response.json()


if __name__ == '__main__':
    unittest.main()
