#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/11/17 9:19
# @Author  : HD
# @File    : EAM_Api_Test.py.py
# @Description :

import unittest

import openpyxl
import requests


# 读取接口测试用例
def read_test_cases_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    test_cases = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True):
        test_case = {
            'name': row[0],
            'method': row[1],
            'url': row[2],
            'params': row[3],
            'expected_result': row[4]
        }
        test_cases.append(test_case)
    return test_cases


class APITestCase(unittest.TestCase):
    def setUp(self):
        '''
        在每个测试方法执行前会执行setUP方法接口基础URL
        '''
        self.base_url = "http://192.168.30.140:8002/eam"
        self.header = {
            'Authorization': 'eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJhZG1pbiIsInN1YiI6Iuezu-e7n-'
                             'euoeeQhuWRmCIsIkNVUlJFTlRfVVNFUl9JTkZPIjoie1wiY29tcGFueUlkXCI6XCJBXCI'
                             'sXCJjb21wYW55TmFtZVwiOlwi6ZuG5Zui5pys57qnXCIsXCJvZmZpY2VJZFwiOlwiQTAyXCIsXCJvZmZp'
                             'Y2VOYW1lXCI6XCLlip7lhazlrqRcIixcInBlcnNvbklkXCI6XCJhZG1pblwiLFwicGVyc29uTmFtZVwiOlwi'
                             '57O757uf566h55CG5ZGYXCIsXCJwZXJzb25Ob1wiOlwiYWRtaW5cIn0iLCJpYXQiOjE3MDA0NDE1MjIsIm'
                             'V4cCI6MTcwMTA0NjMyMn0.fChJgMAjwxx9U7Q6YanItXaEqMgmQgMPN5Eik2DbE1s'
        }

    def test_api(self):
        test_cases = read_test_cases_from_excel("D:\TEST-project\铁投内部系统项目\设备维修维护管理系统\测试文档\EAM接口测试用例.xlsx")
        for test_case in test_cases:
            url = self.base_url + test_case['url']
            method = test_case['method']
            params = test_case['params']
            expected_result = test_case['expected_result']

            # 查看请求接口参数信息
            print(test_case)

        # 初始化response变量
        response = None
        try:
            # 发送请求
            if method == 'GET':
                response = requests.get(url, params=params, headers=self.header)
            elif method == 'POST':
                response = requests.post(url, json=params, headers=self.header)

            # 查看返回结果
            print("===============================分割线======================================")
            print(response.json())

            if response is not None:
                if response.status_code == 400:
                    print(f"接口名称：{url}，跳过接口。")
                    pass
                # 断言--返回状态码为200--->成功
                self.assertEqual(response.status_code, 200)
                # 返回结果符合预期
                self.assertEqual(response.json(), expected_result)
        except Exception as e:
            # 处理异常情况
            print(f"请求出错：{e}")


if __name__ == '__main__':
    unittest.main()
