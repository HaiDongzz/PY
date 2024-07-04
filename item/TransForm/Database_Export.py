import datetime
import time

import cx_Oracle
from openpyxl import Workbook

con = cx_Oracle.connect('t_psmdb/Password@10.11.76.40:1521/T_PSMDB')  # 链接数据库
cur = con.cursor()  # 获取游标

start = time.clock()

now_time = datetime.datetime.now()

time1 = "00:00:00"
time2 = "23:59:59"

for i in range(1, 8, 1):
    day = (now_time + datetime.timedelta(days=-i)).strftime('%Y-%m-%d')
    start_time = day + " " + time
    sql = "select * from t_record where create_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss') order by id desc" % (
        start_time, end_time)

    # 执行sql查询
    cur.execute(sql)
    results = cur.fetchall()

    # 获取行和列
    rows = len(results)
    if len(results):
        cols = len(results[0])  # 判断list是否溢出

    # 创建表格
    wb = Workbook()
    ws = wb.create_sheet('Sheet1', 0)

    # 获取表字段值
    db_title = [i[0] for i in cur.description]
    for i, description in enumerate(db_title):
        ws.cell(row=1, column=1 + i).value = description

    for m in range(rows):
        for n in range(cols):
            ws.cell(row=m + 2, column=n + 1).value = results[m][n]
    wb.save(filename='%s.xlsx' % (day))

cur.close()  # 关闭游标
con.close()  # 关闭链接

end = time.clock()
print('运行时间：%s' % (end - start))
