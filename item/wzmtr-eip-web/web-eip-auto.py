'''
coding:utf-8
@Software:PyCharm
@Time:2023/2/15 14:58
@Author:DuHaidong
'''

'''import requests;
import  json;

url ="http://10.11.82.74:8080/psvServer/index#/lineList"
putcookie=input("cookie输入：")   #不断在变的cookie用录入的方式获取
mycookie =putcookie
#提交所需要的头：包含content-type，user-agent和cookie
myheaders ={"content-type": "application/x-www-form-urlencoded; charset=UTF-8",
          "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.66 Safari/537.36",
          "cookie":mycookie}
#获取到的Form Data 如下：
# id:
# person_type:
# name:
# sex:
# remark:
#所以提交的数据样式应该为{"id": '', "person_type":"", "name": "","sex":, "remark":"" }
#有时候id 或者是 性别可能是 数字类型的也可能是 字符类型的这个可以多尝试下
str = {"StartingStation": '',"TerminalStation":'' , "fare": ''}
#上面的name sex remark是从excel表格中获取到的不是死数据，测试时可以写死
#提交
r = requests.post(url, data=str, headers=myheaders)
#获取返回的结果
print(r.text)'''

# import json;
import openpyxl;
import requests;

wb = openpyxl.load_workbook("data.xlsx")  # 这里不固定的话也可通过控制台录入获取，我写死了
print(wb.active)  # 看当前是哪个工作簿防止选错了
sheet = wb.active  # 就选择当前的工作簿

url = "http://10.11.82.74:8080/psvServer/lineNetworkBased/ticketPriceManage/searchSitePic?staName=&currentPage=10&pageSize=2"
putcookie = input("cookie输入：")  # 不断在变的cookie用录入的方式获取
mycookie = putcookie
# 提交所需要的头：包含content-type，user-agent和cookie
myheaders = {"content-type": "application/x-www-form-urlencoded; charset=UTF-8",
             "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.66 Safari/537.36",
             "cookie": mycookie}


# 这个方法就是循环读取数据然后提交我设置的是10个数据这样可以检查发现错误及时改
def writemsg():
    for i in range(10):
        StartingStation = sheet.cell(row=i + 2, column=1).value  # 第1列 第2行因为i是从0开始的。row是行column是列
        TerminalStation = sheet.cell(row=i + 2, column=4).value  # 这个表格可以自行修改，数据在哪就设置那块
        fare = sheet.cell(row=i + 2, column=5).value
        print(StartingStation, TerminalStation, fare)  # 打印看看
        str = {"StartingStation": '', "TerminalStation": '', "fare": ''}
        r = requests.post(url, data=str, headers=myheaders)
        print(r.text)


if __name__ == '__main__':
    writemsg()
